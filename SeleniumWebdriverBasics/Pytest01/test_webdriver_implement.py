import allure
import openpyxl
import pytest
from allure_commons.types import AttachmentType
from selenium.webdriver.common.by import By
# from SeleniumWebdriverBasics.Tests.loctors import driver
# def get_data():
#     workbook=openpyxl.load_workbook("C:\\Users\\91869\\PycharmProjects\\Selenium Practice\\Pytest01\\newdatafile.xlsx")
#     sheet=workbook["Sheet1"]
#     totalrow=sheet.max_row
#     totalcol=sheet.max_column
#     mainList=[]
#
#     for i in range(2,totalrow+1):
#         dataList=[]
#         for j in range(1,totalcol+1):
#             data=sheet.cell(row=i,column=j).value
#             dataList.insert(j,data)
#         mainList.insert(i,dataList)
#         print(mainList)
#     return mainList
    # return [
    #     ("ABCtest@gmail.com","hdfkdsnfk"),
    #     ("abc@gmail.com","fdsfew123"),
    #     ("xyz@gmail.com","hd3432")
    #
    # ]
@pytest.mark.parametrize("username,password",
                         [
        ("ABCtest@gmail.com","hdfkdsnfk"),
        ("abc@gmail.com","fdsfew123"),
        ("xyz@gmail.com","hd3432")

    ])
def test_dologin(username,password,get_browser):
    driver = get_browser
    driver.find_element(By.ID,"email").send_keys(username)
    driver.find_element(By.ID,"pass").send_keys(password)
    allure.attach(driver.get_screenshot_as_png(),name="dologin",attachment_type=AttachmentType.PNG)
